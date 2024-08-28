from datetime import datetime
from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3
import re
from docxtpl import DocxTemplate
from PyQt5.QtCore import QRegExp, Qt
from PyQt5.QtGui import QRegExpValidator
from PyQt5.QtWidgets import QLineEdit, QLabel, QMessageBox, QTableWidget
import subprocess, os, platform
from Maksim.db_class import DatabaseManager
from collections import ChainMap

class Ui_Client_Add(QtWidgets.QDialog):
    def __init__(self,data):
        super().__init__()
        self.data = data
        self.db_manager = DatabaseManager('My/store_database.db')
        self.alert_msg = QtWidgets.QMessageBox()
        self.alert_msg.setWindowTitle('Information')
        self.con = sqlite3.connect('My/store_database.db')
        self.setObjectName("Client card")
        self.resize(958, 497)
        self.gridLayoutWidget = QtWidgets.QWidget(self)
        self.gridLayoutWidget.setGeometry(QtCore.QRect(30, 20, 901, 461))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        self.gridLayout = QtWidgets.QGridLayout(self.gridLayoutWidget)
        self.gridLayout.setContentsMargins(20, 20, 20, 20)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(self.gridLayoutWidget)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.gridLayout.addWidget(self.tableWidget, 4, 0, 1, 7)
        self.pushButton = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 5, 0, 1, 1)
        self.pushButton_2 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 5, 4, 1, 1)
        self.pushButton_3 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.pushButton_3.setObjectName("pushButton_3")
        self.gridLayout.addWidget(self.pushButton_3, 5, 5, 1, 1)
        self.pushButton_4 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.pushButton_4.setObjectName("pushButton_4")
        self.gridLayout.addWidget(self.pushButton_4, 5, 6, 1, 1)
        self.combo = QtWidgets.QComboBox(self)
        self.gridLayout.addWidget(self.combo, 5, 1, 1, 2)
        self.alert_msg.setWindowTitle('Information')
        self.pushButton.clicked.connect(self.uploadDoc)
        self.pushButton_3.clicked.connect(self.saveClient)
        self.pushButton_4.clicked.connect(self.closeDialog)
        self.pushButton_2.clicked.connect(self.deleteClient)
        self.retranslateUi()
        self.fullTable()


    def retranslateUi(self)->None:
        """В зависимости от того было ли двойное нажатие или нажатие кнопки
        формирует диалоговое окно"""
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Dialog", "Client card"))
        self.pushButton.setText(_translate("Dialog", "upload document"))
        self.pushButton_2.setText(_translate("Dialog", "Delete client"))
        self.pushButton_3.setText(_translate("Dialog", "Save"))
        self.pushButton_4.setText(_translate("Dialog", "cancel"))
        directory = 'My/teamplates'
        name_file = self.get_files(directory)
        self.combo.addItems(name_file)
        self.lineedit_data = []
        with self.con:
            row_col = self.con.execute('Pragma table_info ("Customers")').fetchall()
            self.name_of_col = [i[1] for i in row_col]
            for i in range(len(self.name_of_col)):
                if i == 0 :
                    lineedit = QLineEdit(self)
                    lineedit.setEnabled(False)
                    lineedit.setPlaceholderText('Autofill')
                    if self.data is False:
                        lineedit.setText('')
                        self.pushButton_2.setEnabled(False)
                        self.pushButton.setEnabled(False)
                elif i in (1,2):
                    lineedit = QLineEdit(self)
                    validator = QRegExpValidator(QRegExp('[a-zA-Z]+'))
                    lineedit.setValidator(validator)
                elif i == 4:
                    lineedit = QLineEdit(self)
                    validator = QRegExpValidator(QRegExp('(\s*)?(\+)?([- _():=+]?\d[- _():=+]?){10,14}(\s*)?'))
                    lineedit.setValidator(validator)
                else: lineedit = QLineEdit(self)
                label_name = f'label_{i}'
                label = QLabel(f'{self.name_of_col[i]}')
                lineedit_name = f'lineedit_{i}'
                if self.data is False: lineedit.setText('')
                else :
                    lineedit.setText(self.data[i])
                    self.pushButton_2.setEnabled(True)
                    self.pushButton.setEnabled(True)
                self.lineedit_data.append(lineedit)
                setattr(self, label_name, label)
                setattr(self, lineedit_name, lineedit)
                self.gridLayout.addWidget(lineedit, i+1, 0, 1, 1)
                self.gridLayout.addWidget(label, i+1, 1, 1, 1)
                if 2 <= i <= 3 :
                    self.gridLayout.addWidget(lineedit, i-1, 3, 1, 1)
                    self.gridLayout.addWidget(label, i-1, 4, 1, 1)
                if 4 <= i <= 5:
                    self.gridLayout.addWidget(lineedit, i - 3, 5, 1, 1)
                    self.gridLayout.addWidget(label, i - 3, 6, 1, 1)
                self.setLayout(self.gridLayout)

    def fullTable(self):
        if self.data is not False:
            with self.con:
                base_req = ("SELECT * FROM Transactions_history"
                            " JOIN Customers"
                            " ON Transactions_history.customer_id = Customers.id "
                            f"WHERE Customers.id = {self.data[0]}")
                row_text = self.con.execute(base_req).fetchall() #[ () ]
                respon = self.con.execute("Pragma table_info ('Transactions_history')").fetchall()
                name_of_col = [i[1] for i in respon]
                self.tableWidget.setHorizontalHeaderLabels(name_of_col)
                self.tableWidget.setRowCount(len(row_text))
                self.tableWidget.setColumnCount(len(name_of_col))
                for x in range(len(row_text)):
                    for y in range(len(row_text[x])):
                        item = QtWidgets.QTableWidgetItem(str(row_text[x][y]))
                        item.setFlags(item.flags() & ~ Qt.ItemIsEditable)
                        self.tableWidget.setItem(x, y, item)

    def deleteClient(self)->None:
        """Удаляет по id клиента с подтверждением действительно ли хочет удалить"""
        with self.con:
            try:
                database_query = (f"DELETE FROM Customers WHERE id={self.data[0]}")
                self.alert_msg.setText('Are you sure you want to delete this client?')
                self.alert_msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                res = self.alert_msg.exec_()
                if res == QMessageBox.Ok:
                    self.con.execute(database_query)
                    self.alert_msg.close()
                    QtWidgets.QDialog.close(self)
                else:  self.alert_msg.close()
            except Exception as er:
                print(f'Error : {er}')
                # self.alert_msg.setText(f'Something wrong {er}')
                # self.alert_msg.setStandardButtons(QMessageBox.Ok)
                # self.alert_msg.exec_()

    # def saveClient(self)->None:
    #     """Сохраняет данные нового клиента
    #     Проверяет валидность email"""
    #     line_text = []
    #     pattern = re.compile(r"^\S+@\S+\.\S+$")
    #     is_email = pattern.match(self.lineedit_data[3].text())
    #     if is_email is None:
    #         self.alert_msg.setText('Check email address (user@example.com)')
    #         self.alert_msg.setStandardButtons(QMessageBox.Ok)
    #         self.alert_msg.exec_()
    #     else:
    #         for i in range(len(self.lineedit_data)):
    #             line_text.append(self.lineedit_data[i].text())
    #         with self.con:
    #             val = ', '.join('?' * (len(self.name_of_col)))
    #             column = ', '.join(self.name_of_col)
    #             try:
    #                 query_database = (f"""INSERT OR REPLACE INTO Customers ({column}) VALUES ({val})""")
    #                 self.con.execute(query_database, line_text)
    #                 self.alert_msg.setText('Your data is saved!')
    #                 self.alert_msg.setStandardButtons(QMessageBox.Ok)
    #                 self.alert_msg.exec_()
    #             except Exception as er:
    #                 self.alert_msg.setText(f'Error {er}')
    #                 self.alert_msg.setStandardButtons(QMessageBox.Ok)
    #                 self.alert_msg.exec_()
    #                 print(er)
    #             QtWidgets.QDialog.close(self)

    def saveClient(self) -> None:
        """Сохраняет данные нового клиента и проверяет валидность."""
        line_text = []

        for i in range(len(self.lineedit_data)):
            line_text.append(self.lineedit_data[i].text().strip())
        
        customer_data = dict(zip(self.name_of_col, line_text))
        
        is_valid, validation_message = self.db_manager.validate_data('Customers', customer_data)
        
        if not is_valid:
            self.alert_msg.setText(f"Error: {validation_message}")
            self.alert_msg.setStandardButtons(QMessageBox.Ok)
            self.alert_msg.exec_()
            return
        
        line_text = line_text[1:]
        print("Data being inserted:", line_text)

        with self.con:
            column_names = ', '.join(self.name_of_col[1:])
            placeholders = ', '.join('?' * (len(self.name_of_col) - 1))
            query_database = f"""INSERT INTO Customers ({column_names}) VALUES ({placeholders})"""
            
            try:
                self.con.execute(query_database, line_text)
                self.alert_msg.setText('Your data is saved!')
                self.alert_msg.setStandardButtons(QMessageBox.Ok)
                self.alert_msg.exec_()
            except Exception as er:
                self.alert_msg.setText(f'Error {er}')
                self.alert_msg.setStandardButtons(QMessageBox.Ok)
                self.alert_msg.exec_()
                print(er)
            QtWidgets.QDialog.close(self)


    def forWord(self):
        data_for_doc = self.fill_data()
        doc = DocxTemplate(f'My/teamplates/{self.combo.currentText()}')
        doc.render(data_for_doc)
        doc.save(f'My/doc/{datetime.now().strftime("%Y-%m-%d")}_{self.row_data[1]}_{self.row_data[0]}.docx')


    def forExcel(self):
        file_path = (f'My/teamplates/{self.combo.currentText()}')
        worklist = load_workbook(file_path)
        sheet = worklist.active
        head_in_teamplate = next(sheet.iter_rows(min_row=1,max_row=1,values_only=True))



    def fill_data(self)->dict:
        if self.data is not False:
            with self.con:
                selected_row = self.tableWidget.currentRow()
                if selected_row >= 0:
                    self.row_data = []
                    column_count = self.tableWidget.columnCount()
                    for column in range(column_count):
                        item = self.tableWidget.item(selected_row, column)
                        if item is not None:
                            self.row_data.append(item.text())
                        else:
                            self.row_data.append('')
                customer = self.con.execute(f"SELECT * FROM Customers WHERE id = {self.row_data[7]}").fetchall()
                employee = self.con.execute(f"SELECT * FROM Employees WHERE id = {self.row_data[9]}").fetchall()
                product = self.con.execute(f"SELECT * FROM Products WHERE id = {self.row_data[2]}").fetchall()
                warehouse = self.con.execute(f"SELECT * FROM Warehouses WHERE id = {self.row_data[3]}").fetchall()
                n_customer = self.con.execute("Pragma table_info('Customers')").fetchall()
                col_name_customers = [i[1] for i in n_customer]
                n_employee = self.con.execute("Pragma table_info('Employees')").fetchall()
                col_name_employees = [i[1] for i in n_employee]
                n_product = self.con.execute("Pragma table_info('Products')").fetchall()
                col_name_products = [i[1] for i in n_product]
                n_warehouse = self.con.execute("Pragma table_info('Warehouses')").fetchall()
                col_name_warehouse = [i[1] for i in n_warehouse]
                res = dict(zip(col_name_customers[1:],customer[0][1:]))
                res1 = dict(zip(col_name_employees[1:3],employee[0][1:3]))
                res2 = dict(zip(col_name_products[1:],product[0][1:]))
                res3 = dict(zip(col_name_warehouse[1:],warehouse[0][1:]))
                combina = ChainMap(res, res1, res2, res3)
                return dict(combina)
                # print(customer[0][1:], employee[0][1:3], product[0][1:], warehouse[0][1:])


    def get_files(self,directory):
        return [file for file in os.listdir(directory) if os.path.isfile((os.path.join(directory,file)))]


    def uploadDoc(self)->None:
        """Производит выгрузку документа"""
        if self.combo.currentText().endswith('docx'): self.forWord()
        else: self.forExcel()



        # if platform.system() == 'Darwin': subprocess.call(('open', filepath)) # macOS
        # elif platform.system() == 'Windows': os.startfile(filepath) # Windows
        # else: subprocess.call(('xdg-open', 'My/newDOc.docx')) # linux variants


    def closeDialog(self)->None:
        QtWidgets.QDialog.close(self)


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_Client_Add()
    ui.show()
    sys.exit(app.exec_())
